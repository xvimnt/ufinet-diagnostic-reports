#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import ast
import glob
import os
import sys
from typing import Dict, Tuple, List
import unicodedata
from datetime import datetime
import re

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None  # We'll handle this gracefully and print instructions


def normalize_slug(s: str) -> str:
    if s is None:
        return ""
    # Strip accents and normalize to a simple slug
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join([c for c in s if not unicodedata.combining(c)])  # remove diacritics
    s = s.lower()
    out = []
    prev_us = False
    for ch in s:
        if ch.isalnum():
            out.append(ch)
            prev_us = False
        else:
            if not prev_us:
                out.append("_")
                prev_us = True
    slug = "".join(out).strip("_")
    # Collapse multiple underscores
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug


# Default Spanish -> English slugs mapping (keys/values normalized using normalize_slug)
DEFAULT_MAPPING_RAW = {
    "Temperatura": "temperature",
    "Posible Temperatura": "temperature",
    "Energía Cliente": "energy_client",
    "Energia Cliente": "energy_client",
    "Intermitencia": "intermittency",
    "Puerto LAN": "lan_port",
    "Puerto Lan": "lan_port",
    "Corte de Fibra": "fiber_cut",
    "Corte Fibra": "fiber_cut",
    "Ruta Secundaria Abajo": "secondary_route_down",
    "Ruta Secundaria Down": "secondary_route_down",
    "Posible Corte de Energía": "energy_client",
    "Posible Corte de Energia": "energy_client",
    "Posible Corte / Energía": "energy_client",
    "Posible Falla de Energía": "energy_client",
    "Servicio Ok": "service_ok",
    "No determinado": "undetermined",
    "Suspensión / Baja Lógica": "logical_suspension",
    "SFP Dañado": "sfp_damaged",
    "Sin Información": "no_information",
    "Posible Corte Fibra": "fiber_cut",
    "Posible Corte de Fibra Sin Demarcador": "fiber_cut",
    "Corte en ENNI": "fiber_cut",
    "Posible Corte en ENNI": "fiber_cut",
    "Puerto Inhibido": "port_inhibited",
    "Degradación de Servicio": "service_degradation",
    "Falla Anillo / Bus": "ring_bus_failure",
    "Interfaz Intermitente": "intermittency",
}
DEFAULT_MAPPING: Dict[str, str] = {normalize_slug(k): normalize_slug(v) for k, v in DEFAULT_MAPPING_RAW.items()}


EXPECTED_HEADER_SLUGS = {
    "category",
    "new_result",
    "administrative_code",
    "json",
    "id",
    "created_at",
    "end_at",
}


def infer_fallback_fieldnames(sample_row: List[str]) -> List[str]:
    """Infer sensible column names when CSV files are missing a header row.

    Assumes the canonical order:
    ADMINISTRATIVE_CODE;ID;CATEGORY;JSON;NEW_RESULT;CREATED_AT;END_AT
    and tries to map columns heuristically when more data is present.
    """

    base_names = [
        "ADMINISTRATIVE_CODE",
        "ID",
        "CATEGORY",
        "JSON",
        "NEW_RESULT",
    ]
    if len(sample_row) <= len(base_names):
        names = base_names[: len(sample_row)]
    else:
        extra_count = len(sample_row) - len(base_names)
        names = base_names + [f"EXTRA_{i}" for i in range(1, extra_count + 1)]

    json_idx = None
    new_result_idx = None
    date_labels = ["CREATED_AT", "END_AT"]
    date_idx = 0

    for idx, raw in enumerate(sample_row):
        if idx >= len(names):
            continue
        value = (raw or "").strip()
        lower_value = value.lower()

        if json_idx is None and lower_value.startswith("{") and (
            "'diagnostic_type'" in lower_value or '"diagnostic_type"' in lower_value
        ):
            json_idx = idx

        if new_result_idx is None and (
            "'category'" in lower_value or '"category"' in lower_value
        ):
            new_result_idx = idx

        if names[idx] not in {"ADMINISTRATIVE_CODE", "ID", "CATEGORY", "JSON", "NEW_RESULT"}:
            if re.search(r"\d{4}-\d{2}-\d{2}", value):
                if date_idx < len(date_labels):
                    names[idx] = date_labels[date_idx]
                    date_idx += 1

    if json_idx is not None:
        names[json_idx] = "JSON"

    if new_result_idx is None and names:
        new_result_idx = len(names) - 1
    if new_result_idx is not None:
        names[new_result_idx] = "NEW_RESULT"

    if names.count("JSON") > 1:
        first_json = names.index("JSON")
        for idx in range(first_json + 1, len(names)):
            if names[idx] == "JSON":
                names[idx] = f"EXTRA_{idx}"

    if names.count("NEW_RESULT") > 1:
        first_new_result = names.index("NEW_RESULT")
        for idx in range(first_new_result + 1, len(names)):
            if names[idx] == "NEW_RESULT":
                names[idx] = f"EXTRA_{idx}"

    return names


def parse_new_result_category(value: str) -> str:
    if value is None:
        return ""
    txt = str(value).strip()
    if not txt:
        return ""
    # The field appears to be a Python dict-like string (single quotes), so use ast.literal_eval
    try:
        obj = ast.literal_eval(txt)
        if isinstance(obj, dict):
            return str(obj.get("category", "")).strip()
    except Exception:
        # If parsing fails, try to extract simple pattern 'category': '...'
        # Very naive fallback
        key = "'category':"
        i = txt.find(key)
        if i != -1:
            rest = txt[i + len(key):].strip()
            # Expect something like 'temperature' or "temperature"
            if rest and rest[0] in ("'", '"'):
                q = rest[0]
                try:
                    j = rest.index(q, 1)
                    return rest[1:j]
                except ValueError:
                    pass
    return ""


def evaluate_file(path: str, mapping: Dict[str, str]) -> Tuple[str, int, int, int, float, List[Dict[str, str]]]:
    """
    Returns: (file_name, total, matches, mismatches, match_rate)
    - total counts only rows where both CATEGORY and NEW_RESULT.category were present and parsed
    """
    total = 0
    matches = 0
    mismatch_rows: List[Dict[str, str]] = []

    # Detect delimiter as ';' by default based on provided samples
    delimiter = ';'

    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        peek_reader = csv.reader(f, delimiter=delimiter)
        try:
            first_row = next(peek_reader)
        except StopIteration:
            fname = os.path.basename(path)
            return (fname, 0, 0, 0, 0.0, mismatch_rows)

        has_header = any(
            normalize_slug(cell) in EXPECTED_HEADER_SLUGS for cell in first_row
        )

        f.seek(0)
        if has_header:
            reader = csv.DictReader(f, delimiter=delimiter)
        else:
            fallback_fieldnames = infer_fallback_fieldnames(first_row)
            reader = csv.DictReader(
                f,
                delimiter=delimiter,
                fieldnames=fallback_fieldnames,
            )
        # Normalize header keys to upper for resilience
        field_map = {k: k for k in reader.fieldnames or []}
        # Expected fields
        cat_field = None
        new_res_field = None
        # Optional helpful fields for context in mismatch sheets
        admin_field = None
        id_field = None
        created_at_field = None
        end_at_field = None
        json_field = None
        for k in field_map:
            k_norm = k.strip().upper()
            if k_norm == 'CATEGORY':
                cat_field = field_map[k]
            if k_norm == 'NEW_RESULT':
                new_res_field = field_map[k]
            if k_norm == 'ADMINISTRATIVE_CODE':
                admin_field = field_map[k]
            if k_norm == 'ID':
                id_field = field_map[k]
            if k_norm == 'CREATED_AT':
                created_at_field = field_map[k]
            if k_norm == 'END_AT':
                end_at_field = field_map[k]
            if k_norm == 'JSON':
                json_field = field_map[k]
        if cat_field is None or new_res_field is None:
            # Try a more flexible search
            for k in field_map:
                k_norm = k.strip().lower()
                if 'category' == k_norm:
                    cat_field = field_map[k]
                if 'new_result' == k_norm:
                    new_res_field = field_map[k]
                if 'administrative_code' == k_norm:
                    admin_field = field_map[k]
                if k_norm == 'id':
                    id_field = field_map[k]
                if k_norm == 'created_at':
                    created_at_field = field_map[k]
                if k_norm == 'end_at':
                    end_at_field = field_map[k]
                if k_norm == 'json':
                    json_field = field_map[k]
        if cat_field is None or new_res_field is None:
            # No comparable fields; return zeros
            fname = os.path.basename(path)
            return (fname, 0, 0, 0, 0.0, mismatch_rows)

        for row in reader:
            cat_es_raw = row.get(cat_field, '')
            if cat_es_raw is None:
                cat_es_raw = ''
            cat_es_slug = normalize_slug(cat_es_raw)
            cat_en_raw = parse_new_result_category(row.get(new_res_field, ''))
            cat_en_slug = normalize_slug(cat_en_raw)

            if not cat_es_slug or not cat_en_slug:
                continue  # skip incomplete rows from totals

            total += 1
            mapped = mapping.get(cat_es_slug, None)
            if mapped is None:
                # if mapping missing, assume slug itself might already be english
                mapped = cat_es_slug
            if mapped == cat_en_slug:
                matches += 1
            else:
                # Collect a compact record for the mismatch sheet
                mismatch_rows.append({
                    'ADMINISTRATIVE_CODE': (row.get(admin_field, '') if admin_field else ''),
                    'ID': (row.get(id_field, '') if id_field else ''),
                    'CREATED_AT': (row.get(created_at_field, '') if created_at_field else ''),
                    'END_AT': (row.get(end_at_field, '') if end_at_field else ''),
                    'JSON': (row.get(json_field, '') if json_field else ''),
                    'CATEGORY_ES': str(cat_es_raw),
                    'CATEGORY_EN': str(cat_en_raw),
                })

    mismatches = total - matches
    match_rate = (matches / total) if total else 0.0
    fname = os.path.basename(path)
    return (fname, total, matches, mismatches, match_rate, mismatch_rows)


def _safe_sheet_name(name: str) -> str:
    # Excel sheet names max 31, cannot contain: : \\ / ? * [ ]
    invalid = set(':\\/?*[]')
    clean = ''.join(ch if ch not in invalid else '_' for ch in name)
    if len(clean) > 31:
        clean = clean[:31]
    if not clean:
        clean = 'Sheet'
    return clean


def _save_workbook(wb, out_path: str) -> str:
    """Try saving the workbook to out_path. If the file is locked (e.g., open in Excel),
    save to a timestamped fallback and return the actual written path."""
    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        base, ext = os.path.splitext(out_path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = f"{base}_{ts}{ext}"
        wb.save(alt)
        return alt
    except OSError:
        base, ext = os.path.splitext(out_path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = f"{base}_{ts}{ext}"
        wb.save(alt)
        return alt


def write_excel_report(results: List[Tuple[str, int, int, int, float, List[Dict[str, str]]]], out_path: str) -> str:
    if Workbook is None:
        print("openpyxl is not installed. Please run: pip install -r requirements.txt", file=sys.stderr)
        # Also write a CSV fallback next to the desired xlsx
        csv_path = os.path.splitext(out_path)[0] + ".csv"
        with open(csv_path, 'w', encoding='utf-8', newline='') as f:
            w = csv.writer(f)
            w.writerow(["file_name", "total", "matches", "mismatches", "match_rate"])
            for r in results:
                fname, total, matches, mismatches, rate, _ = r
                w.writerow([fname, total, matches, mismatches, f"{rate:.2%}"])
        print(f"Excel not created (missing openpyxl). Wrote CSV fallback: {csv_path}")
        return csv_path

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["file_name", "total", "matches", "mismatches", "match_rate"])
    for r in results:
        fname, total, matches, mismatches, rate, _ = r
        ws.append([fname, total, matches, mismatches, rate])

    # Format match_rate as percent (column E)
    try:
        from openpyxl.styles import numbers
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
    except Exception:
        pass

    # Add a tab per file with mismatches
    headers = [
        'ADMINISTRATIVE_CODE', 'ID', 'CREATED_AT', 'END_AT', 'JSON',
        'CATEGORY_ES', 'CATEGORY_EN',
    ]
    for r in results:
        fname, _total, _matches, _mismatches, _rate, rows = r
        sheet_name = _safe_sheet_name(os.path.splitext(fname)[0])
        ws_file = wb.create_sheet(title=sheet_name)
        ws_file.append(headers)
        for item in rows:
            ws_file.append([item.get(h, '') for h in headers])

    written_path = _save_workbook(wb, out_path)
    if written_path != out_path:
        print(f"Output file was in use. Saved to fallback: {written_path}")
    return written_path


def main() -> int:
    root = os.path.dirname(os.path.abspath(__file__))
    # Look for CSV inputs in project root and in ./reports/
    search_paths = [os.path.join(root, "*.csv"), os.path.join(root, "reports", "*.csv")]
    found = set()
    for pattern in search_paths:
        for p in glob.glob(pattern):
            # Exclude the generated summary CSV, if any
            if os.path.basename(p).lower() == "category_match_report.csv":
                continue
            found.add(os.path.abspath(p))
    csv_files = sorted(found)
    if not csv_files:
        print("No CSV files found in project root.")
        return 1

    results: List[Tuple[str, int, int, int, float, List[Dict[str, str]]]] = []
    for path in csv_files:
        res = evaluate_file(path, DEFAULT_MAPPING)
        results.append(res)

    out_xlsx = os.path.join(root, "category_match_report.xlsx")
    written_path = write_excel_report(results, out_xlsx)

    # Print a small console summary
    print("Summary (file_name, total, matches, mismatches, match_rate):")
    for r in results:
        print(f"{r[0]}, {r[1]}, {r[2]}, {r[3]}, {r[4]:.2%}")

    print(f"\nReport written to: {written_path}")
    if Workbook is None:
        print("Note: Install dependencies to get an Excel file. See requirements.txt.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
